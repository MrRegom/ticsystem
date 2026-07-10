from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from mantenedores.models import Institucion, Edificio, Piso, Unidad, Vlan
from redes.models import Pma, InfraestructuraRed, RangoIP
from equipos.models import Equipo
from anexos.models import Anexo


HEADER_FILL = PatternFill(start_color="002a54", end_color="002a54", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="ffffff", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="002a54")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="d0d5dd"),
    right=Side(style="thin", color="d0d5dd"),
    top=Side(style="thin", color="d0d5dd"),
    bottom=Side(style="thin", color="d0d5dd"),
)
GROUP_FILL = PatternFill(start_color="eef2ff", end_color="eef2ff", fill_type="solid")


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(len(str(cell.value)))
        best = max(lengths) + 3 if lengths else min_width
        ws.column_dimensions[col_letter].width = min(max(best, min_width), max_width)


def write_title(ws, title, row=1, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 30


def write_rows(ws, start_row, headers, rows):
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))

    for r_idx, row_data in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


class Command(BaseCommand):
    help = "Exporta la infraestructura completa del hospital a un Excel con detalle relacional"

    def add_arguments(self, parser):
        parser.add_argument("--output", "-o", type=str, default="infraestructura_hospital.xlsx",
                            help="Ruta del archivo Excel de salida")

    def handle(self, *args, **options):
        wb = Workbook()

        self._hoja_edificios_pisos(wb)
        self._hoja_pisos_pmas(wb)
        self._hoja_pisos_equipos(wb)
        self._hoja_pisos_anexos(wb)
        self._hoja_pisos_rangos(wb)
        self._hoja_ipam(wb)

        output = options["output"]
        wb.save(output)
        self.stdout.write(self.style.SUCCESS(f"Excel generado: {Path(output).resolve()}"))

    # ------------------------------------------------------------------
    # HOJA 1: Edificios x Pisos (la que pide el usuario)
    # ------------------------------------------------------------------
    def _hoja_edificios_pisos(self, wb):
        ws = wb.active
        ws.title = "Edificios x Pisos"
        write_title(ws, "JERARQUÍA: INSTITUCIÓN → EDIFICIO → PISO", ncols=5)

        headers = ["Institución", "Edificio", "Piso", "Alias", "Activo"]
        rows = []
        for p in Piso.objects.select_related("edificio__institucion").order_by(
            "edificio__institucion__nombre", "edificio__nombre", "nombre"
        ):
            rows.append([
                p.edificio.institucion.nombre if p.edificio.institucion else "(Sin institución)",
                p.edificio.nombre,
                p.nombre,
                p.alias or "",
                "Sí" if p.activo else "No",
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)

    # ------------------------------------------------------------------
    # HOJA 2: Pisos x PMAs
    # ------------------------------------------------------------------
    def _hoja_pisos_pmas(self, wb):
        ws = wb.create_sheet("Pisos x PMAs")
        write_title(ws, "PISOS CON SUS PUNTOS DE INGENIERÍA REAL (PMA)", ncols=7)

        headers = ["Edificio", "Piso", "Código PMA", "Unidad", "Estado", "IPs en el PMA", "Descripción"]
        rows = []
        for pma in Pma.objects.select_related("edificio_piso__edificio", "unidad").order_by(
            "edificio_piso__edificio__nombre", "edificio_piso__nombre", "codigo"
        ):
            rows.append([
                pma.edificio_piso.edificio.nombre,
                pma.edificio_piso.nombre,
                pma.codigo,
                pma.unidad.nombre if pma.unidad else "",
                pma.estado,
                InfraestructuraRed.objects.filter(pma=pma).count(),
                pma.descripcion or "",
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)

    # ------------------------------------------------------------------
    # HOJA 3: Pisos x Equipos
    # ------------------------------------------------------------------
    def _hoja_pisos_equipos(self, wb):
        ws = wb.create_sheet("Pisos x Equipos")
        write_title(ws, "EQUIPOS INVENTARIADOS POR UBICACIÓN (EDIFICIO → PISO → UNIDAD)", ncols=10)

        headers = [
            "Edificio", "Piso", "Unidad",
            "Serial", "Artículo", "Marca", "Modelo",
            "Estado", "IP", "Proveedor",
        ]
        rows = []
        for eq in Equipo.objects.select_related(
            "articulo", "marca", "modelo", "edificio", "piso", "unidad", "estado", "proveedor"
        ).order_by("edificio__nombre", "piso__nombre", "unidad__nombre", "serial_number"):
            rows.append([
                eq.edificio.nombre if eq.edificio else "",
                eq.piso.nombre if eq.piso else "",
                eq.unidad.nombre if eq.unidad else "",
                eq.serial_number,
                eq.articulo.nombre if eq.articulo else "",
                eq.marca.nombre if eq.marca else "",
                eq.modelo.nombre if eq.modelo else "",
                eq.estado.nombre if eq.estado else "",
                str(eq.ip) if eq.ip else "",
                eq.proveedor.nombre if eq.proveedor else "",
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)

    # ------------------------------------------------------------------
    # HOJA 4: Pisos x Anexos
    # ------------------------------------------------------------------
    def _hoja_pisos_anexos(self, wb):
        ws = wb.create_sheet("Pisos x Anexos")
        write_title(ws, "ANEXOS TELEFÓNICOS POR UBICACIÓN", ncols=9)

        headers = [
            "Edificio", "Piso", "Unidad",
            "N° Anexo", "Marca", "Modelo", "Serial",
            "IP", "Estado",
        ]
        rows = []
        for a in Anexo.objects.select_related(
            "edificio", "piso", "unidad"
        ).order_by("edificio__nombre", "piso__nombre", "unidad__nombre", "numero_anexo"):
            rows.append([
                a.edificio.nombre if a.edificio else "",
                a.piso.nombre if a.piso else "",
                a.unidad.nombre if a.unidad else "",
                a.numero_anexo,
                a.marca or "",
                a.modelo or "",
                a.serial_number or "",
                str(a.ip) if a.ip else "",
                a.estado,
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)

    # ------------------------------------------------------------------
    # HOJA 5: Pisos x Rangos IP
    # ------------------------------------------------------------------
    def _hoja_pisos_rangos(self, wb):
        ws = wb.create_sheet("Pisos x Rangos IP")
        write_title(ws, "RANGOS DE IP POR PISO", ncols=9)

        headers = [
            "Edificio", "Piso", "Unidad", "Ubicación",
            "PMA (texto)", "Rack", "Rango", "IP", "Comentario",
        ]
        rows = []
        for r in RangoIP.objects.select_related("piso__edificio").order_by(
            "piso__edificio__nombre", "piso__nombre", "rango", "ip"
        ):
            rows.append([
                r.piso.edificio.nombre,
                r.piso.nombre,
                r.unidad,
                r.ubicacion,
                r.pma,
                r.rack,
                f"{r.rango}.{r.dato}" if r.rango and r.dato else str(r.ip),
                str(r.ip),
                r.comentario or "",
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)

    # ------------------------------------------------------------------
    # HOJA 6: Infraestructura Red (IPAM)
    # ------------------------------------------------------------------
    def _hoja_ipam(self, wb):
        ws = wb.create_sheet("IPAM Completo")
        write_title(ws, "INFRAESTRUCTURA DE RED (IPAM) — TODAS LAS IPs", ncols=13)

        headers = [
            "IP", "Estado", "PMA", "VLAN",
            "Edificio", "Piso", "Unidad",
            "Rack", "Patch Panel", "Switch IP", "Puerto",
            "MAC", "Sector",
        ]
        rows = []
        for ip in InfraestructuraRed.objects.select_related(
            "pma", "vlan", "edificio", "piso", "unidad"
        ).order_by("ip_direccion"):
            rows.append([
                str(ip.ip_direccion), ip.estado,
                ip.pma.codigo if ip.pma else "",
                ip.vlan.nombre if ip.vlan else "",
                ip.edificio.nombre if ip.edificio else "",
                ip.piso.nombre if ip.piso else "",
                ip.unidad.nombre if ip.unidad else "",
                ip.rack or "", ip.patch_panel or "",
                str(ip.switch_ip) if ip.switch_ip else "",
                ip.switch_port or "", ip.mac or "", ip.sector or "",
            ])
        write_rows(ws, 3, headers, rows)
        auto_width(ws)
