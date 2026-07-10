import openpyxl
from django.db import transaction
from mantenedores.models import (
    Articulo, Marca, Modelo, Edificio, Piso, Unidad,
    SistemaOperativo, EstadoEquipo, Proveedor,
    Sector, AreaHospitalaria, Recinto, Institucion, PMA
)
from equipos.models import Equipo

class MargaMargaImporterService:
    @staticmethod
    def _limpiar_texto(valor):
        if valor is None:
            return None
        texto = str(valor).strip()
        return texto if texto and texto.upper() != "NONE" else None

    @staticmethod
    def _obtener_catalogo(modelo, nombre, **kwargs):
        nombre_limpio = MargaMargaImporterService._limpiar_texto(nombre)
        if not nombre_limpio:
            return None
        obj, _ = modelo.objects.get_or_create(nombre=nombre_limpio, defaults=kwargs)
        return obj

    @staticmethod
    @transaction.atomic
    def importar_excel(ruta_archivo):
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        hojas_ignoradas = ['GRAFICO', 'GUARDADOS']

        # Entidades base requeridas
        institucion, _ = Institucion.objects.get_or_create(codigo="HMM", defaults={"nombre": "Hospital Marga Marga"})
        edificio, _ = Edificio.objects.get_or_create(nombre="Edificio Principal", defaults={"institucion": institucion})
        
        articulo, _ = Articulo.objects.get_or_create(nombre="Desconocido (Importación)")
        marca, _ = Marca.objects.get_or_create(nombre="Genérica")
        modelo_default, _ = Modelo.objects.get_or_create(nombre="Genérico", defaults={"marca": marca})
        estado_activo, _ = EstadoEquipo.objects.get_or_create(nombre="En Inventario", defaults={"color_hex": "#17a2b8"})

        total_importados = 0

        for nombre_hoja in wb.sheetnames:
            if any(ignorar in nombre_hoja.upper() for ignorar in hojas_ignoradas):
                continue
            
            # Obtener el Piso desde el nombre de la hoja
            piso_nombre = "Por Definir"
            words = nombre_hoja.upper().split()
            if "PISO" in words:
                try:
                    idx = words.index("PISO")
                    if idx + 1 < len(words):
                        piso_nombre = words[idx+1]
                except ValueError:
                    pass
            if "AUDITORIO" in nombre_hoja.upper():
                piso_nombre = "Auditorio"
                
            piso_obj = MargaMargaImporterService._obtener_catalogo(Piso, piso_nombre, edificio=edificio)

            hoja = wb[nombre_hoja]
            filas = list(hoja.iter_rows(values_only=True))
            if not filas or len(filas) < 2:
                continue

            # Normalizar cabeceras
            cabeceras = [MargaMargaImporterService._limpiar_texto(c) for c in filas[0]]
            
            for fila in filas[1:]:
                if all(val is None for val in fila):
                    continue

                datos_fila = {}
                for i, valor in enumerate(fila):
                    if i < len(cabeceras) and cabeceras[i]:
                        cabecera_limpia = cabeceras[i].upper().replace('\n', '').strip()
                        datos_fila[cabecera_limpia] = valor

                pma_nombre = None
                for k, v in datos_fila.items():
                    if k and "PMA" in k.upper():
                        pma_nombre = MargaMargaImporterService._limpiar_texto(v)
                        break
                correlativo = MargaMargaImporterService._limpiar_texto(datos_fila.get('CORRELATIVO'))
                orden_interno = datos_fila.get('ORDEN INTERNO')
                try:
                    orden_interno = int(orden_interno) if orden_interno else None
                except ValueError:
                    orden_interno = None

                ip = MargaMargaImporterService._limpiar_texto(datos_fila.get('IP'))
                sector_nombre = MargaMargaImporterService._limpiar_texto(datos_fila.get('SECTOR'))
                serie = MargaMargaImporterService._limpiar_texto(datos_fila.get('SERIE'))
                serie_equipo = MargaMargaImporterService._limpiar_texto(datos_fila.get('SERIE EQUIPO'))
                candado = MargaMargaImporterService._limpiar_texto(datos_fila.get('CANDADO/OBSERVACION', datos_fila.get('CANDADO')))
                
                area_nombre = MargaMargaImporterService._limpiar_texto(datos_fila.get('AREA HOSPITALARIA'))
                unidad_nombre = MargaMargaImporterService._limpiar_texto(datos_fila.get('UNIDAD HOSPITALARIA', datos_fila.get('UNIDAD  HOSPITALARIA')))
                recinto_nombre = MargaMargaImporterService._limpiar_texto(datos_fila.get('NOMBRE DE RECINTO', datos_fila.get('NOMBRE DE RECINTO ')))

                if not correlativo and not serie:
                    continue

                # Construir el árbol de la jerarquía
                area_obj = MargaMargaImporterService._obtener_catalogo(AreaHospitalaria, area_nombre) if area_nombre else None
                unidad_obj = MargaMargaImporterService._obtener_catalogo(Unidad, unidad_nombre, area_hospitalaria=area_obj) if unidad_nombre else None
                sector_obj = MargaMargaImporterService._obtener_catalogo(Sector, sector_nombre, piso=piso_obj) if sector_nombre else None
                
                recinto_obj = None
                if recinto_nombre:
                    recinto_obj = MargaMargaImporterService._obtener_catalogo(
                        Recinto, 
                        recinto_nombre, 
                        piso=piso_obj, 
                        sector=sector_obj, 
                        unidad=unidad_obj
                    )

                pma_obj = None
                if pma_nombre and recinto_obj:
                    pma_obj = MargaMargaImporterService._obtener_catalogo(
                        PMA, 
                        pma_nombre, 
                        recinto=recinto_obj
                    )

                serial_final = serie if serie else f"SIN-SERIE-{correlativo or pma_nombre or str(total_importados)}"

                # Crear o actualizar el equipo apuntando SÓLO al PMA (o nulo)
                Equipo.objects.update_or_create(
                    serial_number=serial_final,
                    defaults={
                        "pma": pma_obj,
                        "correlativo": correlativo,
                        "orden_interno": orden_interno,
                        "ip": ip if ip and ip.replace('.', '').isdigit() else None,
                        "serie_corta": serie_equipo,
                        "estado_candado": candado,
                        # Valores por defecto del modelo base
                        "articulo": articulo,
                        "marca": marca,
                        "modelo": modelo_default,
                        "estado": estado_activo,
                    }
                )
                total_importados += 1

        return total_importados
